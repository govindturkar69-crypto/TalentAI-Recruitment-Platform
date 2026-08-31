import os
from contextlib import closing
from io import BytesIO

import openpyxl
from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl.styles import Alignment, Font, PatternFill

from core import get_db_connection, login_required
from services.interview_service import (
    cancel_interview_service,
    complete_interview_service,
    get_recruiter_interviews_for_application,
    schedule_interview_service,
    update_interview_service,
)
from services.recruiter_service import (
    bulk_update_status_service,
    delete_job_service,
    get_application_candidate_profile,
    get_application_resume,
    post_job_service,
    toggle_job_active_service,
    update_job_service,
    update_status_service,
)

recruiter_bp = Blueprint("recruiter", __name__)


@recruiter_bp.route("/recruiter/dashboard")
@login_required(role="recruiter")
def recruiter_dashboard():
    with closing(get_db_connection()) as conn:
        with closing(conn.cursor()) as cur:
            cur.execute(
                "SELECT * FROM jobs WHERE recruiter_id = %s ORDER BY is_active DESC, created_at DESC",
                (session["user_id"],),
            )
            jobs = cur.fetchall()
            cur.execute("SELECT COUNT(*) AS cnt FROM users WHERE role='candidate'")
            total_candidates = cur.fetchone()["cnt"]
            cur.execute("SELECT COUNT(*) AS cnt FROM applications")
            total_applications = cur.fetchone()["cnt"]
            cur.execute("SELECT COUNT(*) AS cnt FROM applications WHERE status='shortlisted'")
            shortlisted = cur.fetchone()["cnt"]

    return render_template(
        "recruiter_dashboard.html",
        jobs=jobs,
        total_candidates=total_candidates,
        total_applications=total_applications,
        shortlisted=shortlisted,
    )


@recruiter_bp.route("/recruiter/post_job", methods=["GET", "POST"])
@login_required(role="recruiter")
def post_job():
    if request.method == "POST":
        title = request.form["job_title"].strip()
        skills = request.form["required_skills"].strip().lower()
        description = request.form["description"].strip()
        location = request.form["location"].strip()
        experience = request.form["experience"].strip()

        if not all([title, skills]):
            flash("Title and required skills are mandatory.", "danger")
            return render_template("post_job.html")

        post_job_service(session["user_id"], title, skills, description, location, experience)

        flash(f"Job '{title}' posted successfully!", "success")
        return redirect(url_for("recruiter.recruiter_dashboard"))

    return render_template("post_job.html")


@recruiter_bp.route("/recruiter/job/<int:job_id>/edit", methods=["GET", "POST"])
@login_required(role="recruiter")
def edit_job(job_id):
    with closing(get_db_connection()) as conn:
        with closing(conn.cursor()) as cur:
            cur.execute("SELECT * FROM jobs WHERE id = %s AND recruiter_id = %s", (job_id, session["user_id"]))
            job = cur.fetchone()

            if not job:
                flash("Job not found.", "danger")
                return redirect(url_for("recruiter.recruiter_dashboard"))

            if request.method == "POST":
                title = request.form["job_title"].strip()
                skills = request.form["required_skills"].strip().lower()
                description = request.form["description"].strip()
                location = request.form["location"].strip()
                experience = request.form["experience"].strip()

                if not all([title, skills]):
                    flash("Title and required skills are mandatory.", "danger")
                    return render_template("edit_job.html", job=job)

                update_job_service(session["user_id"], job_id, title, skills, description, location, experience)
                flash(f"Job '{title}' updated successfully!", "success")
                return redirect(url_for("recruiter.recruiter_dashboard"))

    return render_template("edit_job.html", job=job)


@recruiter_bp.route("/recruiter/job/<int:job_id>/toggle_active", methods=["POST"])
@login_required(role="recruiter")
def toggle_job_active(job_id):
    result = toggle_job_active_service(session["user_id"], job_id)
    flash(result["message"], result["type"])
    return redirect(url_for("recruiter.recruiter_dashboard"))


@recruiter_bp.route("/recruiter/job/<int:job_id>/delete", methods=["POST"])
@login_required(role="recruiter")
def delete_job(job_id):
    result = delete_job_service(session["user_id"], job_id)
    flash(result["message"], result["type"])
    return redirect(url_for("recruiter.recruiter_dashboard"))


@recruiter_bp.route("/recruiter/job/<int:job_id>/applicants")
@login_required(role="recruiter")
def view_applicants(job_id):
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()

    with closing(get_db_connection()) as conn:
        with closing(conn.cursor()) as cur:
            # H3: Verify the recruiter owns this job
            cur.execute("SELECT * FROM jobs WHERE id = %s AND recruiter_id = %s", (job_id, session["user_id"]))
            job = cur.fetchone()
            if not job:
                flash("Job not found.", "danger")
                return redirect(url_for("recruiter.recruiter_dashboard"))

            query = """
                SELECT a.*, u.name AS candidate_name, u.email,
                       a.score, a.matched_skills, a.missing_skills, a.status
                FROM applications a JOIN users u ON a.candidate_id = u.id
                WHERE a.job_id = %s
            """
            params = [job_id]

            if q:
                query += " AND (u.name LIKE %s OR u.email LIKE %s)"
                like_q = f"%{q}%"
                params.extend([like_q, like_q])

            if status and status in ("applied", "shortlisted", "rejected", "hired", "withdrawn"):
                query += " AND a.status = %s"
                params.append(status)

            query += " ORDER BY a.score DESC"

            cur.execute(query, tuple(params))
            applicants = cur.fetchall()

    for i, ap in enumerate(applicants):
        ap["rank"] = i + 1
    return render_template("view_applicants.html", job=job, applicants=applicants, q=q, status=status)


@recruiter_bp.route("/recruiter/applications/bulk_update", methods=["POST"])
@login_required(role="recruiter")
def bulk_update_status():
    app_ids = request.form.getlist("selected_apps")
    new_status = request.form.get("bulk_status", "").strip().lower()
    job_id = request.form.get("job_id")

    if not app_ids or not new_status:
        flash("Please select at least one candidate and a status.", "warning")
        return redirect(url_for("recruiter.view_applicants", job_id=job_id))

    result = bulk_update_status_service(app_ids, new_status, session["user_id"])

    flash(result["message"], result.get("type", "success"))
    return redirect(url_for("recruiter.view_applicants", job_id=job_id))


@recruiter_bp.route("/recruiter/job/<int:job_id>/export")
@login_required(role="recruiter")
def export_applicants(job_id):
    with closing(get_db_connection()) as conn:
        with closing(conn.cursor()) as cur:
            cur.execute("SELECT * FROM jobs WHERE id = %s AND recruiter_id = %s", (job_id, session["user_id"]))
            job = cur.fetchone()
            if not job:
                flash("Job not found.", "danger")
                return redirect(url_for("recruiter.recruiter_dashboard"))

            cur.execute(
                """
                SELECT u.name, u.email, a.score, a.matched_skills, a.missing_skills, a.status, a.applied_at
                FROM applications a JOIN users u ON a.candidate_id = u.id
                WHERE a.job_id = %s ORDER BY a.score DESC
            """,
                (job_id,),
            )
            applicants = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants"
    headers = ["Rank", "Name", "Email", "Score (%)", "Matched Skills", "Missing Skills", "Status", "Applied On"]
    ws.append(headers)
    hfill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_num)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center")
    for i, ap in enumerate(applicants, 1):
        ws.append(
            [
                i,
                ap["name"],
                ap["email"],
                round(ap["score"], 1),
                ap["matched_skills"] or "",
                ap["missing_skills"] or "",
                ap["status"].title(),
                ap["applied_at"].strftime("%d %b %Y") if ap["applied_at"] else "",
            ]
        )
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = min(
            max((len(str(c.value)) if c.value else 0) for c in col_cells) + 4, 50
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    safe = "".join(c if c.isalnum() else "_" for c in job["job_title"])
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"applicants_{safe}.xlsx",
    )


@recruiter_bp.route("/recruiter/application/<int:app_id>/status", methods=["POST"])
@login_required(role="recruiter")
def update_status(app_id):
    new_status = request.form.get("status", "").strip().lower()
    result = update_status_service(app_id, new_status, session["user_id"])
    flash(result["message"], result["type"])
    from app import safe_redirect

    return safe_redirect(url_for("recruiter.recruiter_dashboard"))


@recruiter_bp.route("/recruiter/application/<int:app_id>/candidate")
@login_required(role="recruiter")
def view_candidate_profile(app_id):
    profile_data, err = get_application_candidate_profile(app_id, session["user_id"])
    if err:
        flash(err["error"], err["type"])
        return redirect(url_for("recruiter.recruiter_dashboard"))
    return render_template("recruiter_candidate_profile.html", **profile_data)


@recruiter_bp.route("/recruiter/application/<int:app_id>/resume")
@login_required(role="recruiter")
def view_candidate_resume(app_id):
    filename, err = get_application_resume(app_id, session["user_id"])
    if err:
        flash(err["error"], err["type"])
        return redirect(url_for("recruiter.recruiter_dashboard"))

    if not filename:
        flash("Resume file not found.", "danger")
        from app import safe_redirect

        return safe_redirect(url_for("recruiter.recruiter_dashboard"))

    if filename != os.path.basename(filename):
        flash("Invalid resume file path.", "danger")
        from app import safe_redirect

        return safe_redirect(url_for("recruiter.recruiter_dashboard"))

    if not filename.lower().endswith(".pdf"):
        flash("Invalid resume file format.", "danger")
        from app import safe_redirect

        return safe_redirect(url_for("recruiter.recruiter_dashboard"))

    from pathlib import Path

    upload_dir = Path(current_app.config["UPLOAD_FOLDER"]).resolve()
    candidate_path = (upload_dir / filename).resolve()

    if not candidate_path.is_relative_to(upload_dir):
        flash("Invalid resume file path.", "danger")
        from app import safe_redirect

        return safe_redirect(url_for("recruiter.recruiter_dashboard"))

    if not candidate_path.is_file():
        flash("Resume file not found on server.", "danger")
        from app import safe_redirect

        return safe_redirect(url_for("recruiter.recruiter_dashboard"))

    return send_from_directory(str(upload_dir), filename, mimetype="application/pdf", as_attachment=False)


@recruiter_bp.route("/recruiter/settings")
@login_required(role="recruiter")
def recruiter_settings():
    with closing(get_db_connection()) as conn:
        with closing(conn.cursor()) as cur:
            cur.execute(
                """
                SELECT u.name, u.email, u.company_id,
                       c.name as company_name, c.description, c.website, c.is_active
                FROM users u
                LEFT JOIN companies c ON u.company_id = c.id
                WHERE u.id = %s
            """,
                (session["user_id"],),
            )
            user_data = cur.fetchone()

    return render_template("recruiter_settings.html", user_data=user_data)


# ---------------------------------------------------------------------------
# Interview routes — all authorization comes from services
# ---------------------------------------------------------------------------


@recruiter_bp.route("/recruiter/application/<int:app_id>/interviews")
@login_required(role="recruiter")
def application_interviews(app_id):
    result = get_recruiter_interviews_for_application(app_id, session["user_id"])
    if not result["success"]:
        flash(result["message"], result["type"])
        return redirect(url_for("recruiter.recruiter_dashboard"))
    return render_template(
        "recruiter_interviews.html",
        interviews=result["data"],
        app_info=result["app_info"],
    )


@recruiter_bp.route("/recruiter/application/<int:app_id>/interviews/schedule", methods=["POST"])
@login_required(role="recruiter")
def schedule_interview(app_id):
    scheduled_at = request.form.get("scheduled_at", "").strip()
    duration_minutes = request.form.get("duration_minutes", "").strip()
    mode = request.form.get("mode", "").strip()
    location_or_link = request.form.get("location_or_link", "").strip() or None
    notes = request.form.get("notes", "").strip() or None

    result = schedule_interview_service(
        app_id,
        session["user_id"],
        scheduled_at,
        duration_minutes,
        mode,
        location_or_link,
        notes,
    )
    flash(result["message"], result["type"])
    return redirect(url_for("recruiter.application_interviews", app_id=app_id))


@recruiter_bp.route("/recruiter/interview/<int:interview_id>/update", methods=["POST"])
@login_required(role="recruiter")
def update_interview(interview_id):
    app_id = request.form.get("app_id", type=int)
    scheduled_at = request.form.get("scheduled_at", "").strip()
    duration_minutes = request.form.get("duration_minutes", "").strip()
    mode = request.form.get("mode", "").strip()
    location_or_link = request.form.get("location_or_link", "").strip() or None
    notes = request.form.get("notes", "").strip() or None

    result = update_interview_service(
        interview_id,
        session["user_id"],
        scheduled_at,
        duration_minutes,
        mode,
        location_or_link,
        notes,
    )
    flash(result["message"], result["type"])
    if app_id:
        return redirect(url_for("recruiter.application_interviews", app_id=app_id))
    return redirect(url_for("recruiter.recruiter_dashboard"))


@recruiter_bp.route("/recruiter/interview/<int:interview_id>/cancel", methods=["POST"])
@login_required(role="recruiter")
def cancel_interview(interview_id):
    app_id = request.form.get("app_id", type=int)
    result = cancel_interview_service(interview_id, session["user_id"])
    flash(result["message"], result["type"])
    if app_id:
        return redirect(url_for("recruiter.application_interviews", app_id=app_id))
    return redirect(url_for("recruiter.recruiter_dashboard"))


@recruiter_bp.route("/recruiter/interview/<int:interview_id>/complete", methods=["POST"])
@login_required(role="recruiter")
def complete_interview(interview_id):
    app_id = request.form.get("app_id", type=int)
    result = complete_interview_service(interview_id, session["user_id"])
    flash(result["message"], result["type"])
    if app_id:
        return redirect(url_for("recruiter.application_interviews", app_id=app_id))
    return redirect(url_for("recruiter.recruiter_dashboard"))
