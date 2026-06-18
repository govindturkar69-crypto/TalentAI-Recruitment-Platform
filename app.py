import os
import secrets
from datetime import datetime, timedelta
from io import BytesIO

from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file)
import pymysql
import pymysql.cursors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

load_dotenv()

from models.resume_parser import (
    extract_text_from_pdf, extract_skills,
    get_final_score, score_candidate
)
from analytics.dashboard import (
    get_skill_distribution_chart, get_score_distribution_chart,
    get_job_applicants_chart, get_top_candidates_chart, get_status_chart
)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "fallback_dev_key_change_me")

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
ALLOWED_EXTENSIONS = {"pdf"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_db_connection():
    return pymysql.connect(
        host=os.environ.get("MYSQL_HOST", "localhost"),
        user=os.environ.get("MYSQL_USER", "root"),
        password=os.environ.get("MYSQL_PASSWORD", ""),
        database=os.environ.get("MYSQL_DB", "recruitment_db"),
        cursorclass=pymysql.cursors.DictCursor
    )


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def create_notification(user_id, title, message, notif_type="system"):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO notifications (user_id, title, message, type) VALUES (%s,%s,%s,%s)",
            (user_id, title, message, notif_type)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print("Notification error:", e)


def login_required(role=None):
    from functools import wraps

    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "user_id" not in session:
                flash("Please log in first.", "warning")
                return redirect(url_for("login"))
            if role and session.get("role") != role:
                flash("Access denied.", "danger")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.context_processor
def inject_unread_count():
    if "user_id" not in session:
        return {"unread_count": 0}
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) AS cnt FROM notifications WHERE user_id = %s AND is_read = FALSE",
            (session["user_id"],)
        )
        count = cur.fetchone()["cnt"]
        cur.close()
        conn.close()
        return {"unread_count": count}
    except Exception:
        return {"unread_count": 0}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"].strip()
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        role = request.form["role"]

        if not all([name, email, password, role]):
            flash("All fields are required.", "danger")
            return render_template("register.html")

        hashed = generate_password_hash(password)
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO users (name, email, password, role) VALUES (%s,%s,%s,%s)",
                (name, email, hashed, role)
            )
            conn.commit()
            flash("Registration successful! Please log in.", "success")
            return redirect(url_for("login"))
        except Exception:
            flash("This email is already registered.", "danger")
        finally:
            cur.close()
            conn.close()

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["name"] = user["name"]
            session["role"] = user["role"]
            flash(f"Welcome back, {user['name']}!", "success")

            if user["role"] == "recruiter":
                return redirect(url_for("recruiter_dashboard"))
            return redirect(url_for("candidate_dashboard"))

        flash("Incorrect email or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form["email"].strip().lower()

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()

        if user:
            token = secrets.token_urlsafe(32)
            expires_at = datetime.now() + timedelta(hours=1)
            cur.execute(
                "INSERT INTO password_resets (user_id, token, expires_at) VALUES (%s,%s,%s)",
                (user["id"], token, expires_at)
            )
            conn.commit()
            reset_link = url_for("reset_password", token=token, _external=True)
            flash(f"Password reset link generated. Since email sending isn't configured, "
                  f"use this link now: {reset_link}", "info")
        else:
            flash("If that email is registered, a reset link has been generated.", "info")

        cur.close()
        conn.close()
        return redirect(url_for("forgot_password"))

    return render_template("forgot_password.html")


@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM password_resets WHERE token = %s AND used = FALSE",
        (token,)
    )
    reset_entry = cur.fetchone()

    if not reset_entry or reset_entry["expires_at"] < datetime.now():
        cur.close()
        conn.close()
        flash("This reset link is invalid or has expired. Please request a new one.", "danger")
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        new_password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if new_password != confirm_password:
            flash("Passwords do not match.", "danger")
            cur.close()
            conn.close()
            return render_template("reset_password.html", token=token)

        hashed = generate_password_hash(new_password)
        cur.execute("UPDATE users SET password=%s WHERE id=%s", (hashed, reset_entry["user_id"]))
        cur.execute("UPDATE password_resets SET used=TRUE WHERE id=%s", (reset_entry["id"],))
        conn.commit()
        cur.close()
        conn.close()

        flash("Password reset successful! Please log in with your new password.", "success")
        return redirect(url_for("login"))

    cur.close()
    conn.close()
    return render_template("reset_password.html", token=token)


@app.route("/candidate/dashboard")
@login_required(role="candidate")
def candidate_dashboard():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT a.*, j.job_title, j.location, j.experience
        FROM applications a
        JOIN jobs j ON a.job_id = j.id
        WHERE a.candidate_id = %s
        ORDER BY a.applied_at DESC
    """, (session["user_id"],))
    applications = cur.fetchall()

    cur.execute("SELECT * FROM jobs WHERE is_active = TRUE ORDER BY created_at DESC")
    jobs = cur.fetchall()

    cur.execute(
        "SELECT * FROM resumes WHERE user_id = %s ORDER BY created_at DESC LIMIT 1",
        (session["user_id"],)
    )
    resume = cur.fetchone()

    cur.execute("SELECT job_id FROM saved_jobs WHERE candidate_id = %s", (session["user_id"],))
    saved_job_ids = {row["job_id"] for row in cur.fetchall()}

    cur.close()
    conn.close()

    return render_template("candidate_dashboard.html",
                           applications=applications,
                           jobs=jobs,
                           resume=resume,
                           saved_job_ids=saved_job_ids)


@app.route("/candidate/upload_resume", methods=["GET", "POST"])
@login_required(role="candidate")
def upload_resume():
    if request.method == "POST":
        if "resume" not in request.files:
            flash("Please select a file.", "danger")
            return redirect(request.url)

        file = request.files["resume"]
        if file.filename == "" or not allowed_file(file.filename):
            flash("Only PDF files are allowed.", "danger")
            return redirect(request.url)

        filename = secure_filename(f"user_{session['user_id']}_{file.filename}")
        save_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(save_path)

        raw_text = extract_text_from_pdf(save_path)
        skills = extract_skills(raw_text)
        skills_str = ",".join(skills)

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO resumes (user_id, resume_path, skills, raw_text) VALUES (%s,%s,%s,%s)",
            (session["user_id"], filename, skills_str, raw_text)
        )
        conn.commit()
        cur.close()
        conn.close()

        flash(f"Resume uploaded! {len(skills)} skills found: {skills_str}", "success")
        return redirect(url_for("candidate_dashboard"))

    return render_template("upload_resume.html")


@app.route("/candidate/apply/<int:job_id>", methods=["POST"])
@login_required(role="candidate")
def apply_job(job_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM resumes WHERE user_id = %s ORDER BY created_at DESC LIMIT 1",
        (session["user_id"],)
    )
    resume = cur.fetchone()
    if not resume:
        flash("Please upload your resume first.", "warning")
        cur.close()
        conn.close()
        return redirect(url_for("candidate_dashboard"))

    cur.execute(
        "SELECT id FROM applications WHERE candidate_id=%s AND job_id=%s",
        (session["user_id"], job_id)
    )
    if cur.fetchone():
        flash("You have already applied for this job.", "warning")
        cur.close()
        conn.close()
        return redirect(url_for("candidate_dashboard"))

    cur.execute("SELECT * FROM jobs WHERE id = %s", (job_id,))
    job = cur.fetchone()

    candidate_skills = resume["skills"].split(",") if resume["skills"] else []
    result = get_final_score(
        resume["raw_text"],
        candidate_skills,
        job["required_skills"],
        job["description"] or ""
    )

    matched_str = ",".join(result["matched"])
    missing_str = ",".join(result["missing"])

    cur.execute("""
        INSERT INTO applications
            (candidate_id, job_id, resume_id, score, matched_skills, missing_skills)
        VALUES (%s,%s,%s,%s,%s,%s)
    """, (session["user_id"], job_id, resume["id"],
          result["final_score"], matched_str, missing_str))
    conn.commit()
    cur.close()
    conn.close()

    create_notification(
        session["user_id"],
        "Application Submitted",
        f"You applied for {job['job_title']} with a score of {result['final_score']:.1f}%.",
        "applied"
    )
    create_notification(
        job["recruiter_id"],
        "New Application Received",
        f"{session['name']} applied for {job['job_title']} with a score of {result['final_score']:.1f}%.",
        "applied"
    )

    flash(f"Application submitted! Your score: {result['final_score']:.1f}%", "success")
    return redirect(url_for("candidate_dashboard"))


@app.route("/candidate/withdraw/<int:app_id>", methods=["POST"])
@login_required(role="candidate")
def withdraw_application(app_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT a.*, j.job_title FROM applications a JOIN jobs j ON a.job_id = j.id "
        "WHERE a.id = %s AND a.candidate_id = %s",
        (app_id, session["user_id"])
    )
    app_row = cur.fetchone()

    if not app_row:
        flash("Application not found.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("candidate_dashboard"))

    if app_row["status"] == "hired":
        flash("You can't withdraw an application that has already resulted in a hire.", "warning")
        cur.close()
        conn.close()
        return redirect(url_for("candidate_dashboard"))

    cur.execute("UPDATE applications SET status='withdrawn' WHERE id=%s", (app_id,))
    conn.commit()
    cur.close()
    conn.close()

    flash(f"Application for {app_row['job_title']} has been withdrawn.", "info")
    return redirect(url_for("candidate_dashboard"))


@app.route("/candidate/save_job/<int:job_id>", methods=["POST"])
@login_required(role="candidate")
def save_job(job_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO saved_jobs (candidate_id, job_id) VALUES (%s,%s)",
            (session["user_id"], job_id)
        )
        conn.commit()
        flash("Job saved for later.", "success")
    except Exception:
        flash("This job is already in your saved list.", "info")
    finally:
        cur.close()
        conn.close()
    return redirect(request.referrer or url_for("candidate_dashboard"))


@app.route("/candidate/unsave_job/<int:job_id>", methods=["POST"])
@login_required(role="candidate")
def unsave_job(job_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM saved_jobs WHERE candidate_id=%s AND job_id=%s",
        (session["user_id"], job_id)
    )
    conn.commit()
    cur.close()
    conn.close()
    flash("Job removed from saved list.", "info")
    return redirect(request.referrer or url_for("saved_jobs_page"))


@app.route("/candidate/saved_jobs")
@login_required(role="candidate")
def saved_jobs_page():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT j.*, s.saved_at
        FROM saved_jobs s
        JOIN jobs j ON s.job_id = j.id
        WHERE s.candidate_id = %s
        ORDER BY s.saved_at DESC
    """, (session["user_id"],))
    saved = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("saved_jobs.html", saved_jobs=saved)


@app.route("/recruiter/dashboard")
@login_required(role="recruiter")
def recruiter_dashboard():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM jobs WHERE recruiter_id = %s ORDER BY is_active DESC, created_at DESC",
        (session["user_id"],)
    )
    jobs = cur.fetchall()

    cur.execute("SELECT COUNT(*) AS cnt FROM users WHERE role='candidate'")
    total_candidates = cur.fetchone()["cnt"]

    cur.execute("SELECT COUNT(*) AS cnt FROM applications WHERE candidate_id IN "
                "(SELECT id FROM users WHERE role='candidate')")
    total_applications = cur.fetchone()["cnt"]

    cur.execute("SELECT COUNT(*) AS cnt FROM applications WHERE status='shortlisted'")
    shortlisted = cur.fetchone()["cnt"]

    cur.close()
    conn.close()

    return render_template("recruiter_dashboard.html",
                           jobs=jobs,
                           total_candidates=total_candidates,
                           total_applications=total_applications,
                           shortlisted=shortlisted)


@app.route("/recruiter/post_job", methods=["GET", "POST"])
@login_required(role="recruiter")
def post_job():
    if request.method == "POST":
        title = request.form["job_title"].strip()
        skills = request.form["required_skills"].strip().lower()
        description = request.form["description"].strip()
        location = request.form["location"].strip()
        experience = request.form["experience"].strip()

        if not all([title, skills]):
            flash("Title and required skills are mandatory.", "danger")
            return render_template("post_job.html")

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO jobs (recruiter_id, job_title, required_skills,
                              description, location, experience)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (session["user_id"], title, skills, description, location, experience))
        conn.commit()
        cur.close()
        conn.close()

        flash(f"Job '{title}' posted successfully!", "success")
        return redirect(url_for("recruiter_dashboard"))

    return render_template("post_job.html")


@app.route("/recruiter/job/<int:job_id>/edit", methods=["GET", "POST"])
@login_required(role="recruiter")
def edit_job(job_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM jobs WHERE id = %s AND recruiter_id = %s",
        (job_id, session["user_id"])
    )
    job = cur.fetchone()

    if not job:
        flash("Job not found or you don't have permission to edit it.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("recruiter_dashboard"))

    if request.method == "POST":
        title = request.form["job_title"].strip()
        skills = request.form["required_skills"].strip().lower()
        description = request.form["description"].strip()
        location = request.form["location"].strip()
        experience = request.form["experience"].strip()

        if not all([title, skills]):
            flash("Title and required skills are mandatory.", "danger")
            cur.close()
            conn.close()
            return render_template("edit_job.html", job=job)

        cur.execute("""
            UPDATE jobs
            SET job_title=%s, required_skills=%s, description=%s, location=%s, experience=%s
            WHERE id=%s AND recruiter_id=%s
        """, (title, skills, description, location, experience, job_id, session["user_id"]))
        conn.commit()
        cur.close()
        conn.close()

        flash(f"Job '{title}' updated successfully!", "success")
        return redirect(url_for("recruiter_dashboard"))

    cur.close()
    conn.close()
    return render_template("edit_job.html", job=job)


@app.route("/recruiter/job/<int:job_id>/toggle_active", methods=["POST"])
@login_required(role="recruiter")
def toggle_job_active(job_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT is_active, job_title FROM jobs WHERE id = %s AND recruiter_id = %s",
        (job_id, session["user_id"])
    )
    job = cur.fetchone()

    if not job:
        flash("Job not found.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("recruiter_dashboard"))

    new_state = not job["is_active"]
    cur.execute("UPDATE jobs SET is_active = %s WHERE id = %s", (new_state, job_id))
    conn.commit()
    cur.close()
    conn.close()

    status_word = "reopened" if new_state else "closed"
    flash(f"Job '{job['job_title']}' has been {status_word}.", "success")
    return redirect(url_for("recruiter_dashboard"))


@app.route("/recruiter/job/<int:job_id>/delete", methods=["POST"])
@login_required(role="recruiter")
def delete_job(job_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT job_title FROM jobs WHERE id = %s AND recruiter_id = %s",
        (job_id, session["user_id"])
    )
    job = cur.fetchone()

    if not job:
        flash("Job not found.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("recruiter_dashboard"))

    cur.execute("DELETE FROM jobs WHERE id = %s AND recruiter_id = %s", (job_id, session["user_id"]))
    conn.commit()
    cur.close()
    conn.close()

    flash(f"Job '{job['job_title']}' has been deleted along with all its applications.", "info")
    return redirect(url_for("recruiter_dashboard"))


@app.route("/recruiter/job/<int:job_id>/applicants")
@login_required(role="recruiter")
def view_applicants(job_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM jobs WHERE id = %s", (job_id,))
    job = cur.fetchone()

    cur.execute("""
        SELECT a.*, u.name AS candidate_name, u.email,
               a.score, a.matched_skills, a.missing_skills, a.status
        FROM applications a
        JOIN users u ON a.candidate_id = u.id
        WHERE a.job_id = %s
        ORDER BY a.score DESC
    """, (job_id,))
    applicants = cur.fetchall()
    cur.close()
    conn.close()

    for i, ap in enumerate(applicants):
        ap["rank"] = i + 1

    return render_template("view_applicants.html", job=job, applicants=applicants)


@app.route("/recruiter/applications/bulk_update", methods=["POST"])
@login_required(role="recruiter")
def bulk_update_status():
    app_ids = request.form.getlist("selected_apps")
    new_status = request.form.get("bulk_status")
    job_id = request.form.get("job_id")

    if not app_ids or not new_status:
        flash("Please select at least one candidate and a status.", "warning")
        return redirect(url_for("view_applicants", job_id=job_id))

    conn = get_db_connection()
    cur = conn.cursor()

    for app_id in app_ids:
        cur.execute("UPDATE applications SET status=%s WHERE id=%s", (new_status, app_id))

        cur.execute("""
            SELECT a.candidate_id, j.job_title
            FROM applications a JOIN jobs j ON a.job_id = j.id
            WHERE a.id = %s
        """, (app_id,))
        info = cur.fetchone()

        if info and new_status in ("shortlisted", "rejected", "hired"):
            messages = {
                "shortlisted": f"Good news! You've been shortlisted for {info['job_title']}.",
                "rejected": f"Your application for {info['job_title']} was not selected this time.",
                "hired": f"Congratulations! You've been hired for {info['job_title']}!",
            }
            create_notification(
                info["candidate_id"],
                f"Application {new_status.title()}",
                messages[new_status],
                new_status
            )

    conn.commit()
    cur.close()
    conn.close()

    flash(f"Updated {len(app_ids)} application(s) to '{new_status}'.", "success")
    return redirect(url_for("view_applicants", job_id=job_id))


@app.route("/recruiter/job/<int:job_id>/export")
@login_required(role="recruiter")
def export_applicants(job_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM jobs WHERE id = %s AND recruiter_id = %s", (job_id, session["user_id"]))
    job = cur.fetchone()
    if not job:
        flash("Job not found.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("recruiter_dashboard"))

    cur.execute("""
        SELECT u.name, u.email, a.score, a.matched_skills, a.missing_skills,
               a.status, a.applied_at
        FROM applications a
        JOIN users u ON a.candidate_id = u.id
        WHERE a.job_id = %s
        ORDER BY a.score DESC
    """, (job_id,))
    applicants = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants"

    headers = ["Rank", "Name", "Email", "Score (%)", "Matched Skills",
               "Missing Skills", "Status", "Applied On"]
    ws.append(headers)

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, ap in enumerate(applicants, start=1):
        ws.append([
            i,
            ap["name"],
            ap["email"],
            round(ap["score"], 1),
            ap["matched_skills"] or "",
            ap["missing_skills"] or "",
            ap["status"].title(),
            ap["applied_at"].strftime("%d %b %Y") if ap["applied_at"] else "",
        ])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = "".join(c if c.isalnum() else "_" for c in job["job_title"])
    filename = f"applicants_{safe_title}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/recruiter/application/<int:app_id>/status", methods=["POST"])
@login_required(role="recruiter")
def update_status(app_id):
    new_status = request.form["status"]
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE applications SET status=%s WHERE id=%s", (new_status, app_id))
    conn.commit()

    cur.execute("""
        SELECT a.candidate_id, j.job_title
        FROM applications a JOIN jobs j ON a.job_id = j.id
        WHERE a.id = %s
    """, (app_id,))
    info = cur.fetchone()
    cur.close()
    conn.close()

    if info and new_status in ("shortlisted", "rejected", "hired"):
        messages = {
            "shortlisted": f"Good news! You've been shortlisted for {info['job_title']}.",
            "rejected": f"Your application for {info['job_title']} was not selected this time.",
            "hired": f"Congratulations! You've been hired for {info['job_title']}!",
        }
        create_notification(
            info["candidate_id"],
            f"Application {new_status.title()}",
            messages[new_status],
            new_status
        )

    flash(f"Status updated to: {new_status}", "success")
    return redirect(request.referrer or url_for("recruiter_dashboard"))


@app.route("/recruiter/analytics")
@login_required(role="recruiter")
def analytics():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.*, u.name AS candidate_name, j.job_title
        FROM applications a
        JOIN users u ON a.candidate_id = u.id
        JOIN jobs  j ON a.job_id       = j.id
    """)
    applications = cur.fetchall()
    cur.close()
    conn.close()

    apps_data = [dict(a) for a in applications]

    chart1 = get_skill_distribution_chart(apps_data)
    chart2 = get_score_distribution_chart(apps_data)
    chart3 = get_job_applicants_chart(apps_data)
    chart4 = get_top_candidates_chart(apps_data)
    chart5 = get_status_chart(apps_data)

    stats = {
        "total": len(apps_data),
        "avg_score": round(sum(a.get("score", 0) for a in apps_data) / max(len(apps_data), 1), 1),
        "shortlisted": sum(1 for a in apps_data if a.get("status") == "shortlisted"),
        "hired": sum(1 for a in apps_data if a.get("status") == "hired"),
    }

    return render_template("analytics.html",
                           chart1=chart1, chart2=chart2, chart3=chart3,
                           chart4=chart4, chart5=chart5, stats=stats)


@app.route("/notifications")
@login_required()
def notifications():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM notifications WHERE user_id = %s ORDER BY created_at DESC LIMIT 50",
        (session["user_id"],)
    )
    notifs = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("notifications.html", notifications=notifs)


@app.route("/notifications/mark_all_read", methods=["POST"])
@login_required()
def mark_all_read():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE notifications SET is_read = TRUE WHERE user_id = %s", (session["user_id"],))
    conn.commit()
    cur.close()
    conn.close()
    flash("All notifications marked as read.", "success")
    return redirect(url_for("notifications"))


@app.route("/candidate/profile", methods=["GET", "POST"])
@login_required(role="candidate")
def candidate_profile():
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        bio = request.form.get("bio", "").strip()
        phone = request.form.get("phone", "").strip()
        location = request.form.get("location", "").strip()
        experience_years = request.form.get("experience_years", "").strip()
        linkedin_url = request.form.get("linkedin_url", "").strip()
        github_url = request.form.get("github_url", "").strip()
        portfolio_url = request.form.get("portfolio_url", "").strip()

        cur.execute("SELECT id FROM candidate_profiles WHERE user_id = %s", (session["user_id"],))
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE candidate_profiles
                SET bio=%s, phone=%s, location=%s, experience_years=%s,
                    linkedin_url=%s, github_url=%s, portfolio_url=%s
                WHERE user_id=%s
            """, (bio, phone, location, experience_years,
                  linkedin_url, github_url, portfolio_url, session["user_id"]))
        else:
            cur.execute("""
                INSERT INTO candidate_profiles
                    (user_id, bio, phone, location, experience_years,
                     linkedin_url, github_url, portfolio_url)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (session["user_id"], bio, phone, location, experience_years,
                  linkedin_url, github_url, portfolio_url))

        conn.commit()
        flash("Profile updated successfully!", "success")
        cur.close()
        conn.close()
        return redirect(url_for("candidate_profile"))

    cur.execute("SELECT * FROM candidate_profiles WHERE user_id = %s", (session["user_id"],))
    profile = cur.fetchone()

    cur.execute(
        "SELECT * FROM resumes WHERE user_id = %s ORDER BY created_at DESC LIMIT 1",
        (session["user_id"],)
    )
    resume = cur.fetchone()

    cur.execute(
        "SELECT COUNT(*) AS cnt FROM applications WHERE candidate_id = %s",
        (session["user_id"],)
    )
    application_count = cur.fetchone()["cnt"]
    cur.close()
    conn.close()

    fields_to_check = ["bio", "phone", "location", "experience_years", "linkedin_url"]
    filled = sum(1 for f in fields_to_check if profile and profile.get(f)) if profile else 0
    completion_parts = filled + (1 if resume else 0)
    profile_completion = int((completion_parts / (len(fields_to_check) + 1)) * 100)

    return render_template("candidate_profile.html",
                           profile=profile,
                           resume=resume,
                           application_count=application_count,
                           profile_completion=profile_completion)


@app.route("/candidate/recommendations")
@login_required(role="candidate")
def job_recommendations():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM resumes WHERE user_id = %s ORDER BY created_at DESC LIMIT 1",
        (session["user_id"],)
    )
    resume = cur.fetchone()

    if not resume:
        cur.close()
        conn.close()
        return render_template("job_recommendations.html", resume=None, recommendations=[])

    cur.execute("SELECT * FROM jobs ORDER BY created_at DESC")
    jobs = cur.fetchall()

    cur.execute("SELECT job_id FROM applications WHERE candidate_id = %s", (session["user_id"],))
    applied_job_ids = {row["job_id"] for row in cur.fetchall()}
    cur.close()
    conn.close()

    candidate_skills = resume["skills"].split(",") if resume["skills"] else []

    recommendations = []
    for job in jobs:
        result = score_candidate(candidate_skills, job["required_skills"])
        if result["score"] > 0:
            recommendations.append({
                "job": job,
                "match_score": round(result["score"], 1),
                "matched": result["matched"],
                "already_applied": job["id"] in applied_job_ids,
            })

    recommendations.sort(key=lambda r: r["match_score"], reverse=True)

    return render_template("job_recommendations.html",
                           resume=resume,
                           recommendations=recommendations)


@app.route("/api/jobs")
def api_jobs():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, job_title, required_skills, location, experience FROM jobs")
    jobs = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(jobs)


@app.route("/api/candidate/<int:user_id>/score")
def api_candidate_score(user_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.score, j.job_title
        FROM applications a JOIN jobs j ON a.job_id = j.id
        WHERE a.candidate_id = %s
    """, (user_id,))
    data = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(data)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
